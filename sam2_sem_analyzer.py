import sys
import os
import urllib.request
import numpy as np
import matplotlib
matplotlib.use('TkAgg') # Grįžtame prie TkAgg su teisingais nustatymais
import matplotlib.pyplot as plt
plt.ioff()
from matplotlib.widgets import RectangleSelector
import matplotlib.patheffects as patheffects
import matplotlib.ticker as ticker
import tkinter as tk
from tkinter import simpledialog
import cv2
import pyvista as pv
from skimage.feature import peak_local_max
import torch


# Padėsime atsiųsti modelio svorius su progresu
def download_hook(t):
    last_b = [0]
    def update_to(b=1, bsize=1, tsize=None):
        if tsize is not None:
            t.total = tsize
        t.update((b - last_b[0]) * bsize)
        last_b[0] = b
    return update_to

def ensure_sam2_model():
    # Naudojame SAM 2.1 Hiera Large modelį
    model_path = "sam2.1_hiera_large.pt"
    
    # SAM 2.1 modelio svoriai
    if not os.path.exists(model_path):
        print("SAM 2.1 modelis nerastas! Bandome atsisiųsti naudojant Hugging Face...", flush=True)
        try:
            from huggingface_hub import hf_hub_download
            # Tokenas jau buvo naudotas anksčiau sesijos metu
            hf_hub_download(repo_id='facebook/sam2.1-hiera-large', filename='sam2.1_hiera_large.pt', local_dir='.')
            print("Modelio svoriai sėkmingai paruošti!", flush=True)
        except Exception as e:
            print(f"KLAIDA: Nepavyko automatiškai atsisiųsti modelio: {e}")
            print("Prašome rankiniu būdu įkelti 'sam2.1_hiera_large.pt' į projekto aplanką.")
            sys.exit(1)
    
    return model_path

class SAMAutomaticAnalyzer:
    def __init__(self, image_path):
        self.image_path = image_path
        self.scale = None
        self.roi = None
        self.state = 'SCALE'
        
        # Būsenos sekimui matplotlib lange
        self.scale_p1 = None
        self.scale_p2 = None
        self.scale_dot = None
        self.scale_line = None
        self.roi_selector = None
        
        # 1. Užkrauname nuotrauką
        if not os.path.exists(image_path):
            raise FileNotFoundError(f"Nuotrauka nerasta: {image_path}")

        self.image_bgr = cv2.imread(image_path)
        if self.image_bgr is None:
            raise ValueError(f"Nepavyko užkrauti nuotraukos: {image_path}")
            
        self.orig_rgb = cv2.cvtColor(self.image_bgr, cv2.COLOR_BGR2RGB)
        self.orig_gray = cv2.cvtColor(self.image_bgr, cv2.COLOR_BGR2GRAY)
        
        self.image_rgb = self.orig_rgb.copy()
        self.image_gray = self.orig_gray.copy()
        
        # Pageriname kontrastą mastelio matavimui
        clahe = cv2.createCLAHE(clipLimit=2.0, tileGridSize=(8,8))
        self.enhanced_image = clahe.apply(self.image_gray)
        
        # 2. Inicializuojame SAM 2.1 modelį
        print("Inicializuojamas Segment Anything 2.1 modelis GPU atmintyje...", flush=True)
        try:
            import torch
            from sam2.build_sam import build_sam2
            from sam2.sam2_image_predictor import SAM2ImagePredictor
        except ImportError:
            print("KLAIDA: Nėra įdiegtos 'torch' arba 'sam2' bibliotekos.", flush=True)
            print("Prašome įdiegti: pip install git+https://github.com/facebookresearch/sam2.git", flush=True)
            sys.exit(1)

        model_path = ensure_sam2_model()
        device = "cuda" if torch.cuda.is_available() else "cpu"
        if device == "cpu":
            print("ĮSPĖJIMAS: CUDA nerasta! Pilnas nuotraukos apdorojimas ant CPU užtruks labai ilgai.", flush=True)
        else:
            print(f"CUDA aktyvuota! Naudojamas įrenginys: {torch.cuda.get_device_name(0)}", flush=True)

        # Naudojame build_sam2
        # SAM 2.1 reikalauja konfigūracijos failo pavadinimo (be .yaml) arba pilno kelio
        model_cfg = "configs/sam2.1/sam2.1_hiera_l.yaml"
        
        self.sam_model = build_sam2(
            model_cfg,
            model_path,
            device=device,
            apply_postprocessing=True
        )
        self.device = device
        self.predictor = SAM2ImagePredictor(self.sam_model)
        
        self.attempt_count = 1
        self.masks = []

    def _on_click(self, event):
        """Apdorojami pelės paspaudimai mastelio nustatymui."""
        if self.state != 'SCALE':
            return
        if event.inaxes != self.ax2d:
            return
            
        if self.scale_p1 is None:
            self.scale_p1 = (event.xdata, event.ydata)
            self.scale_dot = self.ax2d.plot(event.xdata, event.ydata, 'ro')[0]
            self.fig2d.canvas.draw_idle()
        elif self.scale_p2 is None:
            self.scale_p2 = (event.xdata, event.ydata)
            self.scale_line = self.ax2d.plot([self.scale_p1[0], self.scale_p2[0]], 
                                             [self.scale_p1[1], self.scale_p2[1]], 'r-', linewidth=2)[0]
            self.fig2d.canvas.draw_idle()
            
            dist_pixels = np.sqrt((self.scale_p2[0] - self.scale_p1[0])**2 + (self.scale_p2[1] - self.scale_p1[1])**2)
            
            root = tk.Tk()
            root.withdraw()
            root.attributes('-topmost', True)
            length_um = simpledialog.askfloat("Mastelis", "Įveskite pažymėtos mastelio linijos ilgį mikrometrais (µm):", minvalue=0.001)
            root.destroy()
            
            if length_um is not None:
                self.scale = length_um / dist_pixels
                print(f"Nustatytas mastelis: 1 pikselis = {self.scale:.5f} µm", flush=True)
                
                # Išvalome mastelio linijas
                self.scale_dot.remove()
                self.scale_line.remove()
                
                # Pereiname į srities (ROI) pasirinkimą
                self.state = 'ROI'
                self.ax2d.set_title(
                    "ŽINGSNIS 2: Pasirinkite apdorojimo sritį!\n"
                    "Apveskite su pele norimą sritį (kad išvengtumėte nereikalingos informacijos apačioje).\n"
                    "Paspauskite 'Enter', kai baigsite.",
                    color='blue', fontweight='bold'
                )
                self.roi_selector = RectangleSelector(
                    self.ax2d, lambda eclick, erelease: None,
                    useblit=True, button=[1], minspanx=5, minspany=5,
                    spancoords='pixels', interactive=True
                )
                self.fig2d.canvas.draw_idle()
            else:
                self.scale_p1 = None
                self.scale_p2 = None
                self.scale_dot.remove()
                self.scale_line.remove()
                self.fig2d.canvas.draw_idle()

    def _on_key(self, event):
        """Apdorojamas Enter klavišas srities patvirtinimui."""
        if event.key == 'enter' and self.state == 'ROI':
            if self.roi_selector and self.roi_selector.extents:
                x1, x2, y1, y2 = [int(v) for v in self.roi_selector.extents]
                xmin, xmax = min(x1, x2), max(x1, x2)
                ymin, ymax = min(y1, y2), max(y1, y2)
                
                # Užtikriname, kad sritis neišeina už nuotraukos ribų
                h, w = self.image_gray.shape
                xmin, xmax = max(0, xmin), min(w, xmax)
                ymin, ymax = max(0, ymin), min(h, ymax)
                
                if xmax - xmin > 10 and ymax - ymin > 10:
                    self.roi = (xmin, xmax, ymin, ymax)
                    print(f"Pasirinkta apdorojimo sritis: X({xmin}-{xmax}), Y({ymin}-{ymax})", flush=True)
                    plt.close(self.fig2d) # Uždaro interaktyvų setup langą
                else:
                    print("Sritis per maža! Pažymėkite didesnę.", flush=True)

    def interactive_setup(self):
        """Atidaro langą, kad vartotojas nustatytų mastelį ir ROI."""
        # Kiekvieną kartą pradedame nuo pilnos nuotraukos (svarbu peržymint sritį)
        self.image_rgb = self.orig_rgb.copy()
        self.image_gray = self.orig_gray.copy()
        
        self.fig2d, self.ax2d = plt.subplots(figsize=(14, 12))
        self.ax2d.imshow(self.enhanced_image, cmap='gray')
        
        from matplotlib.widgets import RectangleSelector
        
        if self.scale is not None:
            # Praleidžiame mastelio nustatymą, einame tiesiai prie ROI
            self.state = 'ROI'
            self.ax2d.set_title(
                "PASIRINKITE APDOROJIMO SRITĮ!\n"
                "Apveskite su pele norimą sritį ir paspauskite 'Enter'.",
                color='blue', fontweight='bold', fontsize=14
            )
            self.roi_selector = RectangleSelector(
                self.ax2d, lambda eclick, erelease: None,
                useblit=True, button=[1], minspanx=5, minspany=5,
                spancoords='pixels', interactive=True
            )
        else:
            self.state = 'SCALE'
            self.scale_p1 = None
            self.scale_p2 = None
            title_text = (
                "ŽINGSNIS 1: Nustatykite mastelį!\n"
                "Spustelkite ant mastelio linijos pradžios, o po to – ant pabaigos."
            )
            self.ax2d.set_title(title_text, fontsize=12, pad=10, color='red', fontweight='bold')
            
        self.ax2d.axis('off')
        self.fig2d.canvas.mpl_connect('button_press_event', self._on_click)
        self.fig2d.canvas.mpl_connect('key_press_event', self._on_key)
        plt.tight_layout()
        plt.show() 

        if self.scale is None:
            print("Mastelis nenustatytas! Naudojamas numatytasis (1 px = 1 µm).", flush=True)
            self.scale = 1.0

        if self.roi is not None:
            xmin, xmax, ymin, ymax = self.roi
            self.image_rgb = self.image_rgb[ymin:ymax, xmin:xmax]
            self.image_gray = self.image_gray[ymin:ymax, xmin:xmax]
            print("Nuotrauka sėkmingai apkarpyta pagal pasirinktą sritį.", flush=True)

    def generate_masks(self):
        """Automatiškai segmentuoja nuotrauką naudojant SAM2AutomaticMaskGenerator."""
        import time
        from sam2.automatic_mask_generator import SAM2AutomaticMaskGenerator
        
        print(f"\n--- VYKDOMAS AUTOMATINIS APTIKIMAS (SAM 2.1, Bandymas #{self.attempt_count}) ---", flush=True)
        print(f"Laikas: {time.strftime('%H:%M:%S')}", flush=True)
        
        # Tankis pagal bandymą: 32 → 48 → 64 taškų tinklelis
        pps = 32 + (self.attempt_count - 1) * 16
        print(f"Tinklelio tankis: {pps} taškų pusėje.", flush=True)
        
        # Sukuriame automatinį generatorių su SAM 2.1 standartiniais parametrais
        mask_generator = SAM2AutomaticMaskGenerator(
            model=self.sam_model,
            points_per_side=pps,
            pred_iou_thresh=0.7,
            stability_score_thresh=0.92,
            stability_score_offset=0.7,
            crop_n_layers=0,
            min_mask_region_area=50,
        )
        
        print("Segmentuojama...", flush=True)
        with torch.no_grad():
            raw_masks = mask_generator.generate(self.image_rgb)
        
        # raw_masks: list of dicts su 'segmentation' (bool [H,W]), 'area', 'predicted_iou' ir kt.
        self.masks = []
        for m in raw_masks:
            seg = m['segmentation'].astype(bool)
            self.masks.append({
                'segmentation': seg,
                'area': int(np.sum(seg)),
                'predicted_iou': float(m.get('predicted_iou', 1.0))
            })
        
        print(f"Detekcija baigta! Rasta {len(self.masks)} objektų.", flush=True)
        self.calculate_statistics(save_files=False)

    def calculate_statistics(self, save_files=True):
        """Apskaičiuoja kiekvieno grūdelio ir bendrą statistiką pagal esamas kaukes."""
        if not self.masks:
            return

        # 3D Gylio (Z) matricos paruošimas
        h, w = self.image_gray.shape
        pixel_area_um2 = self.scale ** 2
        total_roi_area_um2 = h * w * pixel_area_um2
        
        z_scale_factor = (w / 255.0) * 0.1 
        z_um = self.image_gray.astype(float) * z_scale_factor * self.scale
        
        dz_dy, dz_dx = np.gradient(z_um, self.scale, self.scale)
        area_3d_factor = np.sqrt(1 + dz_dx**2 + dz_dy**2)
        
        # Analizės duomenų saugojimas
        stats_list = []
        total_perimeter_um = 0.0
        
        all_boundaries = np.zeros_like(self.image_gray, dtype=bool)
        all_interiors = np.zeros_like(self.image_gray, dtype=bool)
        
        valid_masks = []
        for i, mask_data in enumerate(self.masks):
            mask_bool = mask_data['segmentation'].astype(bool)  # Garantuojame bool tipą
            mask_uint8 = mask_bool.astype(np.uint8) * 255
            
            # 2D Morfologija
            # Perskaičiuojame plotą (ypač po merge/split)
            area_pixels = np.sum(mask_bool)
            mask_data['area'] = int(area_pixels)
            area_um2 = area_pixels * pixel_area_um2
            eq_diameter = 2 * np.sqrt(area_um2 / np.pi)
            if area_um2 < 0.01: continue # Per maži (šiukšlės)
            
            contours, _ = cv2.findContours(mask_uint8, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
            if contours:
                cnt = contours[0]
                perimeter_pixels = cv2.arcLength(cnt, True)
                if perimeter_pixels > 0:
                    sphericity = (4 * np.pi * area_pixels) / (perimeter_pixels**2)
                    sphericity = min(1.0, sphericity)
                else:
                    sphericity = 0
                
                # Filtravimas: jei sferiškumas per mažas (<0.01), tai tikriausiai ne grūdelis, o artefaktas
                if sphericity < 0.01: continue
                
                perimeter_um = perimeter_pixels * self.scale
                
                if len(cnt) >= 5:
                    rect = cv2.minAreaRect(cnt)
                    width, height = rect[1]
                    aspect_ratio = max(width, height) / min(width, height) if width > 0 and height > 0 else 1.0
                else:
                    aspect_ratio = 1.0
            else:
                continue # Nėra kontūro
                
            total_perimeter_um += perimeter_um
            
            # 3D ir Topografija
            area_3d_um2 = np.sum(area_3d_factor[mask_bool]) * pixel_area_um2
            z_mask = z_um[mask_bool]
            mean_z = np.mean(z_mask) if len(z_mask) > 0 else 0
            Ra = np.mean(np.abs(z_mask - mean_z)) if len(z_mask) > 0 else 0
            Rq = np.sqrt(np.mean((z_mask - mean_z)**2)) if len(z_mask) > 0 else 0
            
            # Lūžio topologijai (Intergranuliarinis vs Transgranuliarinis)
            kernel = np.ones((3,3), np.uint8)
            eroded = cv2.erode(mask_uint8, kernel, iterations=1)
            interior_bool = eroded > 0
            boundary_bool = mask_bool & ~interior_bool
            
            all_boundaries |= boundary_bool
            all_interiors |= interior_bool
            
            # Pridedame į mask dictionaries
            mask_data['Area_um2'] = area_um2
            mask_data['eq_diameter'] = eq_diameter
            valid_masks.append(mask_data)
            
            stats_list.append({
                'ID': len(stats_list) + 1,
                'Area_um2': area_um2,
                'Eq_Diameter_um': eq_diameter,
                'Perimeter_um': perimeter_um,
                'Sphericity': sphericity,
                'Aspect_Ratio': aspect_ratio,
                'Area_3D_um2': area_3d_um2,
                'Ra_um': Ra,
                'Rq_um': Rq
            })

        self.masks = valid_masks
        
        # Globalūs skaičiavimai
        grain_boundary_density = (total_perimeter_um / 2) / total_roi_area_um2 if total_roi_area_um2 > 0 else 0
        mean_z_interior = np.mean(z_um[all_interiors]) if np.any(all_interiors) else 0
        mean_z_boundary = np.mean(z_um[all_boundaries]) if np.any(all_boundaries) else 0
        fracture_index = mean_z_interior - mean_z_boundary
        
        # Išsaugome į self kad 3D langas galėtų pasiekti
        self._stats_list = stats_list

        # Spausdiname Globalią Suvestinę
        import datetime
        now = datetime.datetime.now().strftime("%H:%M:%S")
        print(f"\n=== MIKROSTRUKTŪROS STATISTIKA (SAM 2.1 - Atnaujinta {now}) ===", flush=True)
        if len(stats_list) > 0:
            areas = [s['Area_um2'] for s in stats_list]
            diams = [s['Eq_Diameter_um'] for s in stats_list]
            sphericities = [s['Sphericity'] for s in stats_list]
            aspects = [s['Aspect_Ratio'] for s in stats_list]
            roughness_Ra = [s['Ra_um'] for s in stats_list]
            
            fracture_type = ("Stipriai Intergranuliarinis" if fracture_index > 0.5
                             else "Transgranuliarinis" if fracture_index < -0.5
                             else "Mišrus")
            
            self._global_stats = {
                "Algoritmas":          "SAM 2.1 (Segment Anything)",
                "Grūdelių skaičius":   f"{len(stats_list)} vnt.",
                "Vid. diametras":      f"{np.mean(diams):.2f} µm",
                "Min diametras":       f"{np.min(diams):.2f} µm",
                "Max diametras":       f"{np.max(diams):.2f} µm",
                "Vid. plotas":         f"{np.mean(areas):.2f} µm²",
                "Min plotas":          f"{np.min(areas):.2f} µm²",
                "Max plotas":          f"{np.max(areas):.2f} µm²",
                "Vid. anizotropija":   f"{np.mean(aspects):.3f} [be vnt.]",
                "Vid. sferiškumas":    f"{np.mean(sphericities):.3f} [be vnt.]",
                "Ribų tankis":         f"{grain_boundary_density:.4f} µm⁻¹",
                "Vid. Ra":             f"{np.mean(roughness_Ra):.4f} µm",
                "Lūžio indeksas":      f"{fracture_index:.4f} µm",
                "Lūžio tipas":         fracture_type,
            }
            
            for k, v in self._global_stats.items():
                print(f"  {k}: {v}")
        print("=========================================\n", flush=True)

    def visualize_2d_results(self):
        """Atvaizduoja interaktyvų 2D langą rankiniam koregavimui."""
        corrector = InteractiveSAMCorrector(self)
        corrector.show()
        # Po uždarymo, atnaujiname kaukes ir PERSKAIČIUOJAME statistiką
        self.masks = corrector.get_final_masks()
        print(f"Koregavimas baigtas. Perskaičiuojama galutinė statistika...", flush=True)
        self.calculate_statistics(save_files=True)

    def get_final_masks(self):
        return self.masks

    def visualize_3d_pyvista(self):
        """Sukuria 3D vizualizaciją naudojant PyVista su interaktyviomis valdikliais."""
        print("Ruošiamas interaktyvus 3D modelis...", flush=True)
        
        h, w = self.image_gray.shape
        x_grid = np.arange(w)
        y_grid = np.arange(h)[::-1]
        x_grid, y_grid = np.meshgrid(x_grid, y_grid)
        
        z_base = self.image_gray.astype(float)
        z_scale_factor = (w / 255.0) * 0.1
        z_base = z_base * z_scale_factor
        # Cast to float to avoid PyVista warning
        x_grid = x_grid.astype(np.float32)
        y_grid = y_grid.astype(np.float32)

        # Kiekvienam grūdeliui - unikalus normalizuotas indeksas [0..1]
        sorted_masks = sorted(self.masks, key=lambda m: m['area'], reverse=True)
        n = len(sorted_masks)
        grain_colors = np.zeros(self.image_gray.shape, dtype=np.float32)
        for i, mask_data in enumerate(sorted_masks):
            grain_colors[mask_data['segmentation']] = float(i + 1) / max(n, 1)
        
        import pyvista as pv
        import os as _os
        _os.environ.setdefault('QT_API', 'pyqt6')
        import pyvistaqt
        from PyQt6 import QtWidgets as Qw, QtCore
        import sys

        qt_app = Qw.QApplication.instance() or Qw.QApplication(sys.argv)
               
        win = Qw.QMainWindow()
        win.setWindowTitle("SAM 2.1 3D Reljefas - Interaktyvi Analizė")
        win.resize(1380, 1000)

        central = Qw.QWidget()
        main_layout = Qw.QHBoxLayout(central)
        win.setCentralWidget(central)

        # Slenkama valdymo juosta
        ctrl_scroll = Qw.QScrollArea()
        ctrl_scroll.setFixedWidth(280)
        ctrl_scroll.setWidgetResizable(True)
        ctrl_inner = Qw.QWidget()
        ctrl_inner.setStyleSheet("""
            QWidget { background:#f0f0f0; color:#111111; }
            QLabel  { color:#111111; font-size:12px; }
            QComboBox { color:#111111; background:white; }
            QSlider::groove:horizontal { background:#cccccc; height:6px; border-radius:3px; }
            QSlider::handle:horizontal { background:#444; width:14px; height:14px;
                                         margin:-4px 0; border-radius:7px; }
            QPushButton { border-radius:4px; padding:8px; }
        """)
        ctrl_layout = Qw.QVBoxLayout(ctrl_inner)
        ctrl_layout.setAlignment(QtCore.Qt.AlignmentFlag.AlignTop)
        ctrl_layout.setSpacing(4)
        ctrl_scroll.setWidget(ctrl_inner)

        def lbl(t):
            lab = Qw.QLabel(t)
            lab.setStyleSheet("color:#111111; font-weight:bold; margin-top:8px;")
            ctrl_layout.addWidget(lab)

        lbl("<b>Spalvų paletė:</b>")
        palette_combo = Qw.QComboBox()
        PALETTES = ["Set3", "nipy_spectral", "tab20", "hsv", "rainbow", "turbo",
                    "viridis", "plasma", "inferno", "coolwarm", "Paired"]
        palette_combo.addItems(PALETTES)
        palette_combo.setCurrentIndex(0)
        ctrl_layout.addWidget(palette_combo)

        lbl("<b>Glodinimas (iter.):</b>")
        smooth_slider = Qw.QSlider(QtCore.Qt.Orientation.Horizontal)
        smooth_slider.setRange(0, 60)
        smooth_slider.setValue(25)
        smooth_label = Qw.QLabel("25 iter.")
        ctrl_layout.addWidget(smooth_slider)
        ctrl_layout.addWidget(smooth_label)

        lbl("<b>Aukščio mastas:</b>")
        height_slider = Qw.QSlider(QtCore.Qt.Orientation.Horizontal)
        height_slider.setRange(1, 50)
        height_slider.setValue(10)
        height_label = Qw.QLabel("1.0x")
        ctrl_layout.addWidget(height_slider)
        ctrl_layout.addWidget(height_label)

        lbl("<b>Skaidrumas (kuo mažiau – tuo skaidresnis):</b>")
        opacity_slider = Qw.QSlider(QtCore.Qt.Orientation.Horizontal)
        opacity_slider.setRange(0, 90)
        opacity_slider.setValue(50)
        opacity_label = Qw.QLabel("Skaidrumas: 50%")
        ctrl_layout.addWidget(opacity_slider)
        ctrl_layout.addWidget(opacity_label)

        apply_btn = Qw.QPushButton("🔄 Atnaujinti grafiką")
        apply_btn.setStyleSheet("background:#2E7D32;color:white;font-weight:bold;padding:8px;border-radius:4px;")
        ctrl_layout.addWidget(apply_btn)

        # Laukimo pranešimas
        status_lbl = Qw.QLabel("⏳ Perskaičiuojama, prašome palaukti...")
        status_lbl.setStyleSheet("color:#B71C1C;font-weight:bold;padding:6px;background:#FFEBEE;border-radius:4px;")
        status_lbl.setAlignment(QtCore.Qt.AlignmentFlag.AlignCenter)
        status_lbl.setVisible(False)
        ctrl_layout.addWidget(status_lbl)

        export_png_btn = Qw.QPushButton("💾 Išsaugoti PNG")
        export_png_btn.setStyleSheet("background:#1565C0;color:white;font-weight:bold;padding:8px;border-radius:4px;")
        ctrl_layout.addWidget(export_png_btn)

        # Parametrų lentelė
        sep = Qw.QLabel("─────────────────────")
        sep.setStyleSheet("color:#999; margin-top:12px;")
        ctrl_layout.addWidget(sep)
        lbl("<b>Suskaičiuoti parametrai:</b>")

        stats_table = Qw.QTableWidget()
        stats_table.setColumnCount(2)
        stats_table.setHorizontalHeaderLabels(["Parametras", "Reikšmė"])
        stats_table.horizontalHeader().setStretchLastSection(True)
        stats_table.setEditTriggers(Qw.QAbstractItemView.EditTrigger.NoEditTriggers)
        stats_table.setStyleSheet("color:#111; background:white; font-size:11px;")
        stats_table.setMinimumHeight(220)
        stats_table.verticalHeader().setVisible(False)

        # Užpildome iš _global_stats jei yra
        if hasattr(self, '_global_stats') and self._global_stats:
            rows = list(self._global_stats.items())
            stats_table.setRowCount(len(rows))
            for r, (k, v) in enumerate(rows):
                stats_table.setItem(r, 0, Qw.QTableWidgetItem(k))
                stats_table.setItem(r, 1, Qw.QTableWidgetItem(str(v)))
        ctrl_layout.addWidget(stats_table)

        export_xlsx_btn = Qw.QPushButton("📊 Eksportuoti parametrus XLSX")
        export_xlsx_btn.setStyleSheet("background:#6A1B9A;color:white;font-weight:bold;padding:8px;border-radius:4px;")
        ctrl_layout.addWidget(export_xlsx_btn)

        # PyVista ploterėlis
        plotter = pyvistaqt.BackgroundPlotter(show=False)
        pv_widget = plotter.interactor
        main_layout.addWidget(pv_widget, stretch=1)
        main_layout.addWidget(ctrl_scroll)

        def _flash_success(btn, original_style):
            btn.setStyleSheet("background-color: #4CAF50; color: white; font-weight: bold; padding: 8px; border-radius: 4px;")
            # Reikės QTimer, kad grąžintų stilių (bet kadangi jis lokalus, naudojame QtCore.QTimer)
            QtCore.QTimer.singleShot(2000, lambda: btn.setStyleSheet(original_style))

        def _set_busy(busy):
            apply_btn.setEnabled(not busy)
            export_png_btn.setEnabled(not busy)
            export_xlsx_btn.setEnabled(not busy)
            if busy:
                apply_btn.setStyleSheet("background:#888;color:#ccc;font-weight:bold;padding:8px;border-radius:4px;")
                status_lbl.setVisible(True)
            else:
                apply_btn.setStyleSheet("background:#2E7D32;color:white;font-weight:bold;padding:8px;border-radius:4px;")
                status_lbl.setVisible(False)
            qt_app.processEvents()

        def _build_mesh(smooth_iter, height_scale, palette, opacity):
            _set_busy(True)
            plotter.clear()
            z = z_base.astype(np.float32) * (height_scale / 10.0)
            grid = pv.StructuredGrid(x_grid, y_grid, z)
            grid.point_data["Grudelis"] = grain_colors.flatten()
            mesh = grid.extract_surface(algorithm='dataset_surface')
            if smooth_iter > 0:
                mesh = mesh.smooth(n_iter=smooth_iter, relaxation_factor=0.05)
            _flash_success(apply_btn, "background:#2E7D32;color:white;font-weight:bold;padding:8px;border-radius:4px;")
            plotter.add_mesh(mesh, scalars="Grudelis", cmap=palette,
                             show_scalar_bar=False, smooth_shading=True,
                             opacity=opacity / 100.0)
            plotter.remove_all_lights()
            light = pv.Light(
                position=(float(w) * 0.3, float(h) * 1.5, float(z_base.max()) * 6.0),
                focal_point=(float(w) / 2, float(h) / 2, 0.0),
                intensity=0.9
            )
            light.positional = False
            plotter.add_light(light)
            plotter.render()
            _set_busy(False)

        def apply_changes():
            smooth_label.setText(f"{smooth_slider.value()} iter.")
            height_label.setText(f"{height_slider.value()/10:.1f}x")
            actual_opacity = 100 - opacity_slider.value()
            opacity_label.setText(f"Skaidrumas: {actual_opacity}%")
            _build_mesh(smooth_slider.value(), height_slider.value(),
                        palette_combo.currentText(), actual_opacity)

        def export_png():
            _set_busy(True)
            base_path = os.path.splitext(self.image_path)[0]
            out_path = f"{base_path} 3D.png"
            try:
                plotter.screenshot(out_path)
                print(f"3D vaizdas sėkmingai išsaugotas: {out_path}", flush=True)
                _flash_success(export_png_btn, "background:#1565C0;color:white;font-weight:bold;padding:8px;border-radius:4px;")
            except Exception as e:
                print(f"Klaida saugant 3D PNG: {e}", flush=True)
            _set_busy(False)

        def export_xlsx():
            if not hasattr(self, '_stats_list') or not self._stats_list:
                print("Klaida: Nėra duomenų eksportui.", flush=True)
                return
            try:
                import pandas as pd
                import openpyxl
                from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
                from openpyxl.utils import get_column_letter

                base = os.path.splitext(os.path.basename(self.image_path))[0]
                out_path = os.path.join(os.path.dirname(self.image_path),
                                        f"{base} parametrai.xlsx")

                # ── Duomenų paruošimas ──────────────────────────────────
                numeric_cols = ['Area_um2', 'Eq_Diameter_um', 'Perimeter_um',
                                'Sphericity', 'Aspect_Ratio', 'Area_3D_um2',
                                'Ra_um', 'Rq_um']

                df_grains = pd.DataFrame(self._stats_list)
                # Užtikriname stulpelių tvarką
                raw_cols = ['ID'] + [c for c in numeric_cols if c in df_grains.columns]
                df_grains = df_grains[raw_cols]

                # Statistikos eilutės
                stat_rows = []
                for col in numeric_cols:
                    if col in df_grains.columns:
                        v = df_grains[col]
                        stat_rows.append({
                            'Parametras': col,
                            'Vidurkis':        round(float(v.mean()), 4),
                            'Minimum':         round(float(v.min()),  4),
                            'Maximum':         round(float(v.max()),  4),
                            'Std. nuokrypis':  round(float(v.std()),  4),
                        })

                # Globalūs parametrai (be min/max)
                global_rows = []
                if hasattr(self, '_global_stats'):
                    for k, v in self._global_stats.items():
                        global_rows.append({'Parametras': k, 'Reikšmė': str(v)})

                # ── Excel rašymas ───────────────────────────────────────
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "Analizė"

                # Stiliai
                hdr_font   = Font(bold=True, color="FFFFFF", size=10)
                hdr_fill_b = PatternFill("solid", fgColor="1A237E")   # tamsiai mėlyna
                hdr_fill_g = PatternFill("solid", fgColor="1B5E20")   # tamsiai žalia
                hdr_fill_p = PatternFill("solid", fgColor="4A148C")   # violetinė
                row_fill_1 = PatternFill("solid", fgColor="E8F5E9")   # šviesiai žalia
                row_fill_2 = PatternFill("solid", fgColor="E3F2FD")   # šviesiai mėlyna
                glob_fill  = PatternFill("solid", fgColor="FFF9C4")   # geltona
                center     = Alignment(horizontal="center", vertical="center")
                left_al    = Alignment(horizontal="left",   vertical="center")
                thin       = Side(style="thin", color="BBBBBB")
                border     = Border(left=thin, right=thin, top=thin, bottom=thin)

                def _cell(r, c, val, font=None, fill=None, align=None, num_fmt=None):
                    cell = ws.cell(row=r, column=c, value=val)
                    if font:  cell.font      = font
                    if fill:  cell.fill      = fill
                    if align: cell.alignment = align
                    if num_fmt: cell.number_format = num_fmt
                    cell.border = border
                    return cell

                # ══════════════════════════════════════════════════════
                # KAIRĖ PUSĖ — individualūs grūdelių duomenys (A..I)
                # ══════════════════════════════════════════════════════
                # Antraštė
                for ci, col_name in enumerate(raw_cols, start=1):
                    _cell(1, ci, col_name, font=hdr_font, fill=hdr_fill_b, align=center)
                    ws.column_dimensions[get_column_letter(ci)].width = 16

                # Duomenys
                for ri, row_dict in enumerate(self._stats_list, start=2):
                    fill = row_fill_1 if ri % 2 == 0 else None
                    for ci, col_name in enumerate(raw_cols, start=1):
                        val = row_dict.get(col_name, "")
                        _cell(ri, ci, val,
                              fill=fill,
                              align=center,
                              num_fmt="0.0000" if col_name != "ID" else "0")

                # ══════════════════════════════════════════════════════
                # DEŠINĖ PUSĖ — statistikos lentelė (prasideda stulp. K = 11)
                # ══════════════════════════════════════════════════════
                GAP   = 1           # tarpinis tuščias stulpelis
                SCOL  = len(raw_cols) + GAP + 1   # pradžios stulpelis
                GCOL  = SCOL + 5 + GAP            # Globalių parametrų stulpelis (po 5 stat stulpelių + gap)

                stat_headers = ["Parametras", "Vidurkis", "Minimum",
                                "Maximum", "Std. nuokrypis"]

                # Viename lape: kairėje raw, per vidurį statistika, dešinėje globalūs
                SROW = 1

                # Antraštės (Statistika)
                for ci, h in enumerate(stat_headers, start=SCOL):
                    _cell(SROW, ci, h, font=hdr_font, fill=hdr_fill_g, align=center)
                    ws.column_dimensions[get_column_letter(ci)].width = 20

                # Statistikos eilutės
                for ri, sr in enumerate(stat_rows, start=SROW + 1):
                    fill = row_fill_1 if ri % 2 == 0 else row_fill_2
                    _cell(ri, SCOL,     sr['Parametras'],     fill=fill, align=left_al)
                    _cell(ri, SCOL + 1, sr['Vidurkis'],       fill=fill, align=center, num_fmt="0.0000")
                    _cell(ri, SCOL + 2, sr['Minimum'],        fill=fill, align=center, num_fmt="0.0000")
                    _cell(ri, SCOL + 3, sr['Maximum'],        fill=fill, align=center, num_fmt="0.0000")
                    _cell(ri, SCOL + 4, sr['Std. nuokrypis'], fill=fill, align=center, num_fmt="0.0000")

                # Globalūs parametrai — "Grūdelių skaičius" ir kt. (GRETŽIA STATISTIKAI)
                # Antraštė
                _cell(SROW, GCOL,     "Parametras", font=hdr_font, fill=hdr_fill_p, align=center)
                _cell(SROW, GCOL + 1, "Reikšmė",   font=hdr_font, fill=hdr_fill_p, align=center)
                ws.column_dimensions[get_column_letter(GCOL)].width = 25
                ws.column_dimensions[get_column_letter(GCOL + 1)].width = 20

                for ri, gr in enumerate(global_rows, start=SROW + 1):
                    _cell(ri, GCOL,     gr['Parametras'], fill=glob_fill, align=left_al)
                    _cell(ri, GCOL + 1, gr['Reikšmė'],   fill=glob_fill, align=center)

                # Įšaldome 1 eilutę ir ID stulpelį
                ws.freeze_panes = "B2"

                wb.save(out_path)
                print(f"Statistika sėkmingai išsaugota: {out_path}", flush=True)
                _flash_success(export_xlsx_btn, "background:#6A1B9A;color:white;font-weight:bold;padding:8px;border-radius:4px;")
            except Exception as e:
                print(f"Eksporto klaida: {e}", flush=True)

        apply_btn.clicked.connect(apply_changes)
        export_png_btn.clicked.connect(export_png)
        export_xlsx_btn.clicked.connect(export_xlsx)
        smooth_slider.valueChanged.connect(lambda v: smooth_label.setText(f"{v} iter."))
        height_slider.valueChanged.connect(lambda v: height_label.setText(f"{v/10:.1f}x"))
        opacity_slider.valueChanged.connect(lambda v: opacity_label.setText(f"Skaidrumas: {100 - v}%"))

        _build_mesh(25, 10, "Set3", 50)
        win.show()
        print("Atidaromas 3D langas! Jį galite sukioti su pele, priartinti su ratuku.", flush=True)
        qt_app.exec()

    def run(self):
        while True:
            # 1. Pasiruošimas (Mastelis ir ROI)
            self.interactive_setup()
            
            while True:
                # Svarbu: nunuliname prašymus kiekvieno ciklo pradžioje
                self.restart_requested = False
                self.roi_restart_requested = False
                
                # 2. Kaukių generavimas
                self.generate_masks()
                
                # 3. 2D Rezultatų rodymas
                self.visualize_2d_results()
                
                if self.roi_restart_requested:
                    # Vartotojas nori iš naujo žymėti sritį (ROI)
                    self.roi = None # Išvalome ROI
                    # Reikia grįžti į interactive_setup (išorinį ciklą)
                    break
                
                if not self.restart_requested:
                    # Vartotojas nepaspaudė "Pakartoti aptikimą", baigiame šią nuotrauką
                    break
                
                # Jei restart_requested == True, ciklas sukasi vėl (tik generate_masks)
                self.attempt_count += 1

            if not self.roi_restart_requested:
                # Jei nebuvo prašymo peržymėti ROI, galime pereiti prie 3D
                self.visualize_3d_pyvista()
                break
            # Jei buvo roi_restart_requested, išorinis ciklas suksis vėl nuo interactive_setup

class InteractiveSAMCorrector:
    def __init__(self, analyzer):
        self.analyzer = analyzer
        self.masks = list(analyzer.masks)
        self.image_rgb = analyzer.image_rgb
        self.selected_indices = []
        
        self.fig, self.ax = plt.subplots(figsize=(14, 12))
        self.fig.canvas.manager.set_window_title("SAM 3.1 Rankinis Koregavimas")
        
        self.cid_click = self.fig.canvas.mpl_connect('button_press_event', self.on_click)
        self.cid_key = self.fig.canvas.mpl_connect('key_press_event', self.on_key)
        
        self.base_img_obj = None
        
        # Pridedame mygtukus
        from matplotlib.widgets import Button
        
        # Sukuriame papildomas ašis mygtukams (pos: [left, bottom, width, height])
        ax_repeat = self.fig.add_axes([0.02, 0.01, 0.15, 0.04])
        self.btn_repeat = Button(ax_repeat, '🔄 Pakartoti aptikimą', color='#E3F2FD', hovercolor='#BBDEFB')
        self.btn_repeat.on_clicked(self.on_repeat_clicked)
        
        ax_roi = self.fig.add_axes([0.18, 0.01, 0.15, 0.04])
        self.btn_roi = Button(ax_roi, '✂️ Iš naujo sritį', color='#FFF9C4', hovercolor='#FFF176')
        self.btn_roi.on_clicked(self.on_roi_clicked)
        
        ax_export = self.fig.add_axes([0.34, 0.01, 0.15, 0.04])
        self.btn_export = Button(ax_export, '📷 PNG eksportas', color='#E8F5E9', hovercolor='#C8E6C9')
        self.btn_export.on_clicked(self.on_export_clicked)
        
        self.render()

    def on_repeat_clicked(self, event):
        print("Vartotojas paprašė pakartoti aptikimą...", flush=True)
        self.analyzer.restart_requested = True
        plt.close(self.fig)

    def on_roi_clicked(self, event):
        print("Vartotojas paprašė iš naujo pažymėti sritį...", flush=True)
        self.analyzer.roi_restart_requested = True
        plt.close(self.fig)

    def on_export_clicked(self, event):
        base_path = os.path.splitext(self.analyzer.image_path)[0]
        export_path = f"{base_path} grudeliu atpazynimas.png"
        
        # Laikinai paslepiame mygtukų ašis ir apatinį užrašą, kad jie nepatektų į nuotrauką
        self.btn_repeat.ax.set_visible(False)
        self.btn_roi.ax.set_visible(False)
        self.btn_export.ax.set_visible(False)
        if hasattr(self, 'info_text_obj'):
            self.info_text_obj.set_visible(False)
        
        try:
            self.fig.savefig(export_path, dpi=300, bbox_inches='tight')
            print(f"Vizualizacija sėkmingai išsaugota: {export_path}", flush=True)
            
            # Pažymime mygtuką žaliai kaip sėkmės ženklą
            self.btn_export.ax.set_facecolor('#4CAF50') 
            self.btn_export.label.set_text("✅ IŠSAUGOTA")
            self.fig.canvas.draw()
            
            # Po sekundės grąžiname pradinį tekstą (spalvą paliekame žalią iki kito veiksmo)
            plt.pause(1.5)
            self.btn_export.label.set_text("📷 PNG eksportas")
            self.fig.canvas.draw()
            
        except Exception as e:
            print(f"Klaida saugant PNG: {e}", flush=True)
        
        # Grąžiname viską atgal
        self.btn_repeat.ax.set_visible(True)
        self.btn_roi.ax.set_visible(True)
        self.btn_export.ax.set_visible(True)
        if hasattr(self, 'info_text_obj'):
            self.info_text_obj.set_visible(True)
        self.fig.canvas.draw()

    def render(self):
        try:
            # Išvalome tik tekstus
            for t in list(self.ax.texts): t.remove()
            
            h, w = self.analyzer.image_gray.shape
            display_img = self.image_rgb.copy().astype(np.float32) / 255.0
            
            for i, mask_data in enumerate(self.masks):
                mask_bool = mask_data['segmentation']
                if i in self.selected_indices:
                    color = np.array([1.0, 1.0, 0])
                    alpha = 0.6
                else:
                    # Naudojame bandymo numerį seed'ui, kad spalvos šiek tiek skirtųsi pakartojus
                    np.random.seed(i + getattr(self.analyzer, 'attempt_count', 0) * 100)
                    color = np.random.random(3)
                    alpha = 0.4
                
                display_img[mask_bool] = (1 - alpha) * display_img[mask_bool] + alpha * color
                
                y, x = np.where(mask_bool)
                if len(y) > 0:
                    # Naudojame jau suskaičiuotą Area_um2 arba paskaičiuojame iš pikselių
                    area_um2 = mask_data.get('Area_um2', mask_data.get('area', 0) * (self.analyzer.scale ** 2))
                    label = f"{i+1}.\n{area_um2:.1f} µm²"
                    self.ax.text(np.mean(x), np.mean(y), label, 
                                color='white', fontsize=8, fontweight='bold', ha='center', va='center',
                                path_effects=[patheffects.withStroke(linewidth=2, foreground='black')])

            if self.base_img_obj is None:
                self.base_img_obj = self.ax.imshow(display_img)
                self.ax.axis('off')
            else:
                self.base_img_obj.set_data(display_img)
            
            self.ax.set_title(f"Apdorotoje srityje rasta grūdelių: {len(self.masks)}", 
                              fontsize=13, fontweight='bold', color='#1A237E', pad=15)
            
            # Visa informacija apačioje
            full_info = "RANKINIS APDOROJIMAS  |  [J] Sujungti  |  [D] Ištrinti  |  [S] Perskirti  |  [ESC] Baigti"
            
            self.info_text_obj = self.ax.text(0.5, -0.02, full_info, transform=self.ax.transAxes, 
                         ha='center', va='top', fontsize=11, fontweight='bold', 
                         color='#B71C1C', bbox=dict(facecolor='#F5F5F5', alpha=0.9, edgecolor='#B71C1C', boxstyle='round,pad=0.5'))
            
            self.fig.canvas.draw_idle()
        except Exception as e:
            print(f"Atvaizdavimo klaida: {e}", flush=True)

    def on_click(self, event):
        if event.inaxes != self.ax: return
        
        ix, iy = event.xdata, event.ydata
        if ix is None or iy is None: return
        
        ix, iy = int(ix), int(iy)
        h, w = self.analyzer.image_gray.shape
        if not (0 <= ix < w and 0 <= iy < h): return
        
        found_idx = -1
        for i, mask_data in enumerate(self.masks):
            if mask_data['segmentation'][iy, ix]:
                found_idx = i
                break
        
        if found_idx != -1:
            if found_idx in self.selected_indices:
                self.selected_indices.remove(found_idx)
            else:
                self.selected_indices.append(found_idx)
            self.render()

    def on_key(self, event):
        k = event.key.lower()
        
        if k == 'j': 
            self.merge_selected()
        elif k == 'd': 
            self.delete_selected()
        elif k == 's': 
            if len(self.selected_indices) == 1: self.split_selected()
        elif k == 'escape':
            plt.close(self.fig)

    def merge_selected(self):
        if len(self.selected_indices) < 2: 
            print("Pasirinkite bent 2 grūdelius sujungimui!", flush=True)
            return
        print(f"Sujungiami grūdeliai: {[i+1 for i in self.selected_indices]}", flush=True)
        new_mask_bool = np.zeros_like(self.masks[0]['segmentation'], dtype=bool)
        total_area = 0
        for idx in self.selected_indices:
            new_mask_bool |= self.masks[idx]['segmentation']
            total_area += self.masks[idx]['area']
        
        new_mask = self.masks[self.selected_indices[0]].copy()
        new_mask['segmentation'] = new_mask_bool
        new_mask['area'] = total_area
        
        for idx in sorted(self.selected_indices, reverse=True):
            self.masks.pop(idx)
        self.masks.append(new_mask)
        self.selected_indices = []
        self.render()

    def delete_selected(self):
        if not self.selected_indices: return
        print(f"Ištrinami grūdeliai: {[i+1 for i in self.selected_indices]}", flush=True)
        for idx in sorted(self.selected_indices, reverse=True):
            self.masks.pop(idx)
        self.selected_indices = []
        self.render()

    def split_selected(self):
        idx = self.selected_indices[0]
        mask_bool = self.masks[idx]['segmentation']
        print(f"Perskiriamas grūdelis {idx+1}. Pažymėkite du taškus centre naujų dalių.", flush=True)
        pts = plt.ginput(2, timeout=0)
        if len(pts) < 2: return
        
        # Naudojame analyzer jau turimą prediktorių
        predictor = self.analyzer.predictor
        predictor.set_image(self.analyzer.image_rgb)
        
        new_masks = []
        for pt in pts:
            m, scores, _ = predictor.predict(
                point_coords=np.array([pt]),
                point_labels=np.array([1]),
                multimask_output=True
            )
            best_m = m[np.argmax(scores)]
            best_m &= mask_bool
            if np.sum(best_m) > 10:
                new_masks.append({
                    'segmentation': best_m,
                    'area': int(np.sum(best_m)),
                    'predicted_iou': 0.99
                })
        
        if new_masks:
            self.masks.pop(idx)
            self.masks.extend(new_masks)
        
        self.selected_indices = []
        self.render()

    def show(self):
        print("Atidaromas interaktyvus koregavimo langas...", flush=True)
        # Priverčiame matplotlib fokusą ir blokavimą
        self.fig.canvas.draw()
        plt.show(block=True)

    def get_final_masks(self):
        return self.masks


if __name__ == "__main__":
    import sys
    import tkinter as tk
    from tkinter import filedialog
    
    if len(sys.argv) > 1:
        IMAGE_PATH = sys.argv[1]
    else:
        root = tk.Tk()
        root.withdraw()
        root.attributes('-topmost', True)
        IMAGE_PATH = filedialog.askopenfilename(
            title="Pasirinkite SEM nuotrauką analizei",
            filetypes=[("Nuotraukos", "*.tif *.tiff *.png *.jpg *.jpeg")]
        )
        root.destroy()
        
    if IMAGE_PATH and os.path.exists(IMAGE_PATH):
        try:
            app = SAMAutomaticAnalyzer(IMAGE_PATH)
            app.run()
        except Exception as e:
            import traceback
            traceback.print_exc()
            print(f"Klaida paleidžiant aplikaciją: {e}", flush=True)
            input("Paspauskite Enter, kad uždarytumėte...")

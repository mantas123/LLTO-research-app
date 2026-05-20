# Copyright 2026 Mantas Jonas Marcinkevičius
#
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
#     http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the limitations under the License.

import os
import numpy as np
import matplotlib.pyplot as plt
import matplotlib.ticker as ticker
from matplotlib.figure import Figure
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg, NavigationToolbar2Tk
from matplotlib.widgets import RectangleSelector
import tkinter as tk
from tkinter import ttk, messagebox

from language_driver import _, get_config_val

def to_sci_unicode(value, decimals=2):
    """Konvertuoja skaičių į 10ⁿ formatą su Unicode laipsniais ir kableliu."""
    if abs(value) < 1e-25: return "0"
    s = ("{:.%dE}" % decimals).format(value)
    base, exp = s.split('E')
    base = base.replace('.', ',')
    exp_int = int(exp)
    superscripts = {'0':'⁰','1':'¹','2':'²','3':'³','4':'⁴','5':'⁵','6':'⁶','7':'⁷','8':'⁸','9':'⁹','-':'⁻','+':''}
    unicode_exp = "".join(superscripts.get(c, c) for c in str(exp_int))
    return f"{base}·10{unicode_exp}"

def show_arc_width(app):
    # Gauti visas prieinamas temperatūras iš projekto
    str_to_orig = {}
    all_avail_temps = []
    for k in app.vars.keys():
        s = f"{k:g}" if isinstance(k, (int, float)) else str(k)
        str_to_orig[s] = k
        all_avail_temps.append(s)
        
    # Rūšiuojame pagal skaitinę vertę
    try:
        all_avail_temps.sort(key=lambda x: float(x.replace(',', '.')))
    except:
        all_avail_temps.sort()
        
    if not all_avail_temps:
        messagebox.showwarning(_('msg_warning', 'Warning'), _('arc_no_data', 'Please load data first!'))
        return
        
    # Pažymėtos temperatūros pagrindiniame lange
    selected = sorted([t for t, v in app.vars.items() if v.get()], 
                      key=lambda x: float(str(x).replace(',', '.')) if isinstance(x, (int, float)) else float(str(x).replace(',', '.')))
    
    # Nustatome pradinį pasirinkimą ir rodomas temperatūras
    if selected:
        initial_temp = _('arc_all_selected', 'All Selected')
        plot_temps = [str_to_orig[f"{t:g}" if isinstance(t, (int, float)) else str(t)] for t in selected]
    else:
        initial_temp = all_avail_temps[0]
        plot_temps = [str_to_orig[initial_temp]]
        
    # Toplevel langas
    sw = tk.Toplevel(app.root)
    sw.title(_('arc_window_title', 'CeraMIS - Arc Analysis (Rectangle Selection)'))
    w, h = app.center_window(sw, 1700, 1350)
    
    # Pagrindinis rėmelis horizontaliam išdėstymui
    main_frame = tk.Frame(sw)
    main_frame.pack(fill=tk.BOTH, expand=True)
    
    # Kairysis rėmelis grafikui
    plot_frame = tk.Frame(main_frame)
    
    # 1. Pirmiausia pakuojame dešinįjį rėmelį (Sidebar), kad jis niekada nepasislėptų ir nesusispaustų
    sidebar = tk.Frame(main_frame, width=340, bg="#F5F5F5", bd=1, relief="solid")
    sidebar.pack(side=tk.RIGHT, fill=tk.Y, padx=0, pady=0)
    sidebar.pack_propagate(False) # Užfiksuojame plotį
    
    # 2. Tada pakuojame kairįjį rėmelį grafikui, kuris užims visą likusią laisvą erdvę
    plot_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
    
    # Antraštės sidebare
    tk.Label(sidebar, text=_('arc_sidebar_title', 'ARC WIDTH ANALYSIS'), font=('Arial', 12, 'bold'), bg="#F5F5F5", fg="#1A237E").pack(pady=15)
    
    desc_text = _('arc_desc_usage', (
        "Usage:\n"
        "1. Right-click and drag a rectangle around the desired arc.\n"
        "2. Click [Lock Arc] or press 'S' to save the arc.\n"
        "3. Drag around another arc and repeat."
    ))
    tk.Label(sidebar, text=desc_text, font=('Arial', 9), justify="left", bg="#F5F5F5", fg="#333", wraplength=310).pack(pady=5, padx=10)
    
    # Temperatūros parinkimas (Dropdown)
    tk.Label(sidebar, text=_('arc_temp_selection', 'Temperature Selection:'), font=('Arial', 10, 'bold'), bg="#F5F5F5").pack(anchor="w", padx=15, pady=(10, 2))
    
    combo_values = [_('arc_all_selected', 'All Selected')] + all_avail_temps
    
    temp_var = tk.StringVar(value=initial_temp)
    temp_combo = ttk.Combobox(sidebar, textvariable=temp_var, values=combo_values, state="readonly", font=('Arial', 10))
    temp_combo.pack(fill="x", padx=15, pady=2)
    
    # Mygtukų konteineris
    btn_frame = tk.Frame(sidebar, bg="#F5F5F5")
    btn_frame.pack(pady=15, padx=10, fill="x")
    
    # Užfiksuotų rezultatų Text laukas
    tk.Label(sidebar, text=_('arc_locked_results_title', 'Locked Results:'), font=('Arial', 10, 'bold'), bg="#F5F5F5").pack(anchor="w", padx=15, pady=(10, 5))
    
    # Scrollable text widget
    txt_frame = tk.Frame(sidebar)
    txt_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)
    
    txt_scroll = tk.Scrollbar(txt_frame)
    txt_scroll.pack(side=tk.RIGHT, fill=tk.Y)
    
    res_text = tk.Text(txt_frame, height=20, width=35, font=('Consolas', 9), yscrollcommand=txt_scroll.set)
    res_text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
    txt_scroll.config(command=res_text.yview)
    
    res_text.insert(tk.END, _('arc_no_locked_arcs', 'No locked arcs yet.\n'))
    res_text.config(state=tk.DISABLED)
    
    # Duomenų rinkiniai ir Matplotlib ašis
    dpi = 100
    fig = Figure(figsize=((w - 340) / dpi, h / dpi), dpi=dpi, facecolor='white')
    ax = fig.add_subplot(111)
    
    all_datasets = {}
    cmap = plt.cm.viridis
    
    app.selector_markers = []
    app.locked_arc_artists = []
    app.locked_arc_data = []
    app.current_arc_result = None

    # 1. Pirmiausia aprašome on_select
    def on_select(eclick, erelease):
        # Valome tik laikinai nubrėžtus markerius
        for m in app.selector_markers:
            try: m.remove()
            except: pass
        app.selector_markers.clear()
        app.current_arc_result = None
        
        x1, y1 = eclick.xdata, eclick.ydata
        x2, y2 = erelease.xdata, erelease.ydata
        xmin, xmax = min(x1, x2), max(x1, x2)
        ymin, ymax = min(y1, y2), max(y1, y2)
        
        results_text = []
        temp_current_fit = []
        
        for temp, (f_data, z_n, color) in all_datasets.items():
            mask = (z_n.real >= xmin) & (z_n.real <= xmax) & \
                   (-z_n.imag >= ymin) & (-z_n.imag <= ymax)
            
            if np.any(mask):
                sel_f = f_data[mask]
                sel_zr = z_n.real[mask]
                sel_zi = -z_n.imag[mask]
                
                m = ax.plot(sel_zr, sel_zi, 'o', color=color, mec='red', mew=1.5, ms=6, alpha=0.8)[0]
                app.selector_markers.append(m)
                
                r_simple = np.max(sel_zr) - np.min(sel_zr)
                p_idx = np.argmax(sel_zi)
                fp = sel_f[p_idx]
                
                # Aproksimuojame lanką apskritimu (Least Squares)
                if len(sel_zr) >= 3:
                    A = np.c_[2 * sel_zr, 2 * sel_zi, np.ones_like(sel_zr)]
                    B = sel_zr**2 + sel_zi**2
                    C_sol, _resid, _rank, _s = np.linalg.lstsq(A, B, rcond=None)
                    xc, yc = C_sol[0], C_sol[1]
                    R_circ2 = C_sol[2] + xc**2 + yc**2
                    
                    if R_circ2 > yc**2:
                        dx = np.sqrt(R_circ2 - yc**2)
                        x_int1, x_int2 = xc - dx, xc + dx
                        r_fit = x_int2 - x_int1
                        phi = np.arcsin(min(1.0, abs(yc) / np.sqrt(R_circ2)))
                        n_cpe = 1.0 - (2.0 * phi / np.pi)
                        
                        # Nupiešiame laikiną lanką
                        x_arc = np.linspace(x_int1, x_int2, 100)
                        y_arc = yc + np.sqrt(np.clip(R_circ2 - (x_arc - xc)**2, 0, None))
                        arc_line = ax.plot(x_arc, y_arc, '--', color=color, lw=2)[0]
                        int_pts = ax.plot([x_int1, x_int2], [0, 0], 'x', color='black', ms=8, mew=2, zorder=5)[0]
                        app.selector_markers.extend([arc_line, int_pts])
                        
                        if fp > 0 and r_fit > 0:
                            omega_p = 2 * np.pi * fp
                            q_cpe = 1 / (r_fit * (omega_p ** n_cpe))
                            c_eq = 1 / (omega_p * r_fit)
                            
                            # Skaičiuojame Chi-kvadratą (CNLS paklaidų nuokrypis)
                            w_j = 2 * np.pi * sel_f
                            z_model = 1.0 / (1.0 / r_fit + q_cpe * (1j * w_j)**n_cpe)
                            z_exp = sel_zr - 1j * sel_zi
                            residuals = (np.abs(z_exp - z_model) ** 2) / (np.abs(z_exp) ** 2 + 1e-20)
                            chi_square = float(np.sum(residuals))
                        else:
                            q_cpe, c_eq, chi_square = 0, 0, 0.0
                            
                        res_str = (f"{temp}K: R={to_sci_unicode(r_fit)} Ω·m, n={n_cpe:.3f}, Q={to_sci_unicode(q_cpe)}\n"
                                   f"   C_eq={to_sci_unicode(c_eq)} F/m, Chi²={to_sci_unicode(chi_square)}\n"
                                   f"   Int:[{to_sci_unicode(x_int1)}, {to_sci_unicode(x_int2)}]")
                        results_text.append(res_str)
                        
                        temp_current_fit.append({
                            'temp': temp, 'r_fit': r_fit, 'n_cpe': n_cpe, 'q_cpe': q_cpe, 'c_eq': c_eq,
                            'x_int1': x_int1, 'x_int2': x_int2, 'color': color,
                            'fit_x': x_arc, 'fit_y': y_arc, 'chi_square': chi_square
                        })
                    else:
                        c_simple = 1 / (2 * np.pi * fp * r_simple) if fp > 0 and r_simple > 0 else 0
                        results_text.append(f"{temp}K: Max-Min R={to_sci_unicode(r_simple)} Ω·m, C={to_sci_unicode(c_simple)} F/m")
                else:
                    c_simple = 1 / (2 * np.pi * fp * r_simple) if fp > 0 and r_simple > 0 else 0
                    results_text.append(f"{temp}K: Max-Min R={to_sci_unicode(r_simple)} Ω·m, C={to_sci_unicode(c_simple)} F/m")
        
        if not results_text:
            ax.set_title(_('arc_no_points_selected', 'No points in selected region!'))
        else:
            display_limit = 6
            display_text = _('arc_results_label', 'Results:\n') + "\n".join(results_text[:display_limit])
            if len(results_text) > display_limit:
                display_text += _('arc_and_more_label', '\n... and {} more').format(len(results_text)-display_limit)
            ax.set_title(display_text, fontsize=10, fontweight='bold')
            
            # Saugome dabartinio žymėjimo duomenis, kad galėtume juos užfiksuoti
            app.current_arc_result = {
                'temp_fits': temp_current_fit,
                # Nukopijuojame gautus markerius, kad perneštume į locked
                'markers': list(app.selector_markers)
            }
        
        fig.canvas.draw()

    # 2. Antra, aprašome redraw_plot (kuris dabar taip pat iš naujo inicializuoja RectangleSelector)
    def redraw_plot(plot_temps):
        ax.clear()
        ax.set_xlabel(_('axis_real_z_norm', "Z', Ω·m"))
        ax.set_ylabel(_('axis_imag_z_norm', "-Z'', Ω·m"))
        ax.grid(True, linestyle='--', alpha=0.6)
        
        ax.xaxis.set_major_formatter(ticker.ScalarFormatter(useMathText=True))
        ax.yaxis.set_major_formatter(ticker.ScalarFormatter(useMathText=True))
        ax.ticklabel_format(style='sci', scilimits=(-3, 4), axis='both')
        ax.set_aspect('auto')
        
        all_datasets.clear()
        n_sel = len(plot_temps)
        k_LA, k_AL = app._get_geometric_factors()
        
        for i, orig_temp in enumerate(plot_temps):
            f, z = app.get_filtered_data(orig_temp)
            if len(z) < 2: continue
            
            z_n = z * k_AL
            color = cmap(i / max(n_sel - 1, 1)) if n_sel > 1 else '#1f77b4'
            lbl = f"{orig_temp:g}K" if isinstance(orig_temp, (int, float)) else f"{orig_temp}K"
            ax.plot(z_n.real, -z_n.imag, 'o-', markersize=6, color=color, alpha=0.7, label=lbl)
            all_datasets[orig_temp] = (f, z_n, color)

        if len(all_datasets) <= 15:
            ax.legend(fontsize=8)
            
        # Iš naujo nupiešiame visus užfiksuotus lankus, kurie priklauso rodomoms temperatūroms
        app.locked_arc_artists.clear()
        for idx, item in enumerate(app.locked_arc_data):
            for fit in item['temp_fits']:
                matching_orig = None
                for plotted_temp in all_datasets.keys():
                    s1 = f"{plotted_temp:g}" if isinstance(plotted_temp, (int, float)) else str(plotted_temp)
                    s2 = f"{fit['temp']:g}" if isinstance(fit['temp'], (int, float)) else str(fit['temp'])
                    if s1 == s2:
                        matching_orig = plotted_temp
                        break
                        
                if matching_orig is not None:
                    color = all_datasets[matching_orig][2]
                    arc_line = ax.plot(fit['fit_x'], fit['fit_y'], '-', color=color, lw=2.5)[0]
                    int_pts = ax.plot([fit['x_int1'], fit['x_int2']], [0, 0], 'x', color='black', ms=8, mew=2, zorder=5)[0]
                    app.locked_arc_artists.extend([arc_line, int_pts])
                    
        ax.set_title(_('arc_plot_instruction', 'Use right click to drag a rectangle around an arc\nLocked arcs: {}').format(len(app.locked_arc_data)))
        
        # Privalome iš naujo atkurti RectangleSelector po ax.clear()!
        app.rs = RectangleSelector(ax, on_select,
                                     useblit=False,
                                     button=[1, 3], 
                                     minspanx=0, minspany=0,
                                     interactive=True,
                                     props=dict(facecolor='red', edgecolor='red', alpha=0.15, fill=True))
        fig.canvas.draw()

    # Užkrauname pradinį vaizdą ir sugeneruojame RectangleSelector
    redraw_plot(plot_temps)

    def on_temp_change(event):
        selected_val = temp_var.get()
        if selected_val == _('arc_all_selected', 'All Selected'):
            active_selected = [t for t, v in app.vars.items() if v.get()]
            plot_temps = [str_to_orig[f"{t:g}" if isinstance(t, (int, float)) else str(t)] for t in active_selected]
        else:
            plot_temps = [str_to_orig[selected_val]]
            
        # Valome laikiną žymėjimą prieš keičiant temperatūrą
        for m in app.selector_markers:
            try: m.remove()
            except: pass
        app.selector_markers.clear()
        app.current_arc_result = None
        
        redraw_plot(plot_temps)

    temp_combo.bind("<<ComboboxSelected>>", on_temp_change)

    def update_sidebar_text():
        res_text.config(state=tk.NORMAL)
        res_text.delete("1.0", tk.END)
        if not app.locked_arc_data:
            res_text.insert(tk.END, _('arc_no_locked_arcs', 'No locked arcs yet.\n'))
        else:
            for idx, item in enumerate(app.locked_arc_data):
                for fit in item['temp_fits']:
                    lbl = f"{fit['temp']:g}" if isinstance(fit['temp'], (int, float)) else str(fit['temp'])
                    res_text.insert(tk.END, _('arc_sidebar_arc_info', 'Arc #{idx} ({temp} K):\n').format(idx=idx+1, temp=lbl))
                    res_text.insert(tk.END, f"  R = {to_sci_unicode(fit['r_fit'])} Ω·m\n")
                    res_text.insert(tk.END, f"  n = {fit['n_cpe']:.3f}\n")
                    res_text.insert(tk.END, f"  Q = {to_sci_unicode(fit['q_cpe'])}\n")
                    res_text.insert(tk.END, f"  C_eq = {to_sci_unicode(fit['c_eq'])} F/m\n")
                    res_text.insert(tk.END, f"  Chi² = {to_sci_unicode(fit.get('chi_square', 0.0))}\n\n")
        res_text.config(state=tk.DISABLED)
        res_text.see(tk.END)

    def lock_current_arc():
        if not app.current_arc_result:
            messagebox.showwarning(_('msg_warning', 'Warning'), _('arc_please_select_first', 'Please drag a rectangle around an arc first!'))
            return
        
        # Pernešame laikinai nupieštus elementus į permanentinius užfiksuotus
        # Pakeičiame jų stilių į ištisinę liniją (solid)
        for m in app.current_arc_result['markers']:
            if hasattr(m, 'get_linestyle') and m.get_linestyle() == '--':
                m.set_linestyle('-')
                m.set_linewidth(2.5) # Storesnė, kad išsiskirtų
            app.locked_arc_artists.append(m)
        
        # Pašaliname iš selector_markers sąrašo, kad kitas RectangleSelector jų neištrintų
        app.selector_markers.clear()
        
        # Pridedame duomenis į užfiksuotų duomenų bazę
        app.locked_arc_data.append(app.current_arc_result)
        app.current_arc_result = None
        
        update_sidebar_text()
        ax.set_title(_('arc_locked_success_title', 'Arc successfully locked! Select another one.\nLocked arcs: {}').format(len(app.locked_arc_data)))
        fig.canvas.draw()

    def clear_locked_arcs():
        # Pašaliname visus nupieštus locked lanko markerius iš matplotlib ašies
        for m in app.locked_arc_artists:
            try: m.remove()
            except: pass
        app.locked_arc_artists.clear()
        
        # Išvalome ir selector markerius
        for m in app.selector_markers:
            try: m.remove()
            except: pass
        app.selector_markers.clear()
        
        app.locked_arc_data.clear()
        app.current_arc_result = None
        
        update_sidebar_text()
    def export_locked_arcs_to_excel():
        if not app.locked_arc_data:
            messagebox.showwarning(_('msg_warning', 'Warning'), _('arc_no_arcs_export', 'No locked arcs to export!'))
            return
        
        from tkinter import filedialog
        file_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel Files", "*.xlsx")],
            initialfile="CeraMIS_Lanku_Rezultatai.xlsx",
            title=_('arc_export_save_title', 'Save Arc Analysis Results')
        )
        if not file_path: return
        
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = _('arc_excel_sheet_title', 'Arc Results')
            
            headers = [
                _('arc_col_index', "Arc Index"), 
                _('arc_col_temp', "Temperature"), 
                _('arc_col_resistance', "Resistance R"), 
                _('arc_col_n', "CPE Exponent n"), 
                _('arc_col_q', "CPE Capacitance Q"), 
                _('arc_col_ceq', "Equivalent Capacitance C_eq"), 
                _('arc_col_x1', "Intercept X1"), 
                _('arc_col_x2', "Intercept X2"),
                _('arc_col_chi2', "Chi-Square (Chi²)")
            ]
            units = [
                "", 
                "K", 
                "\u2126\u00b7m", 
                "", 
                "F\u00b7s^(n-1)/m", 
                "F/m", 
                "\u2126\u00b7m", 
                "\u2126\u00b7m",
                ""
            ]
            
            # Stiliai
            hdr_font = Font(name='Arial', size=11, bold=True, color="FFFFFF")
            unit_font = Font(name='Arial', size=10, italic=True, color="555555")
            hdr_fill = PatternFill("solid", fgColor="1F497D") # Graži tamsiai mėlyna
            unit_fill = PatternFill("solid", fgColor="E9EDF4") # Švelni melsva/pilkšva
            
            center_align = Alignment(horizontal="center", vertical="center")
            thin_side = Side(style="thin", color="CCCCCC")
            cell_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
            
            # Įrašome antraštes (Row 1 - Origin Long Name)
            for col_idx, h in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx, value=h)
                cell.font = hdr_font
                cell.fill = hdr_fill
                cell.alignment = center_align
                cell.border = cell_border
                
            # Įrašome vienetus (Row 2 - Origin Units)
            for col_idx, u in enumerate(units, 1):
                cell = ws.cell(row=2, column=col_idx, value=u)
                cell.font = unit_font
                cell.fill = unit_fill
                cell.alignment = center_align
                cell.border = cell_border
                
            # Įrašome duomenis (nuo Row 3)
            row_idx = 3
            for arc_idx, item in enumerate(app.locked_arc_data):
                for fit in item['temp_fits']:
                    data = [
                        _('arc_item_prefix', "Arc #{}").format(arc_idx+1),
                        fit['temp'],
                        fit['r_fit'],
                        fit['n_cpe'],
                        fit['q_cpe'],
                        fit['c_eq'],
                        fit['x_int1'],
                        fit['x_int2'],
                        fit.get('chi_square', 0.0)
                    ]
                    for col_idx, val in enumerate(data, 1):
                        cell = ws.cell(row=row_idx, column=col_idx)
                        cell.alignment = center_align
                        cell.border = cell_border
                        
                        # Skaitinių verčių formatavimas Exceliui ir teisingas tipavimas
                        if col_idx >= 2:
                            try:
                                cell.value = float(val)
                                if col_idx in [3, 5, 6, 7, 8, 9]:
                                    cell.number_format = '0.00E+00'
                                elif col_idx == 4:
                                    cell.number_format = '0.000'
                                else:
                                    cell.number_format = '0.0'
                            except:
                                cell.value = val
                        else:
                            cell.value = val
                    row_idx += 1
                    
            # Stulpelių pločių nustatymas
            for col in ws.columns:
                max_len = max(len(str(cell.value or '')) for cell in col)
                col_letter = openpyxl.utils.get_column_letter(col[0].column)
                ws.column_dimensions[col_letter].width = max(max_len + 4, 12)
                
            wb.save(file_path)
            messagebox.showinfo(_('msg_success', 'Success'), _('arc_export_success', 'Origin-compatible data successfully exported to:\n{}').format(file_path))
        except Exception as e:
            messagebox.showerror(_('msg_error', 'Error'), _('arc_export_error', 'Failed to export to Excel:\n{}').format(e))
 
    def import_locked_arcs_from_excel():
        from tkinter import filedialog
        import os
        init_file = get_config_val('default_spectrum_file', '')
        init_dir = None
        if init_file and os.path.exists(init_file):
            init_dir = os.path.dirname(init_file)
        elif init_file and os.path.isdir(init_file):
            init_dir = init_file
            
        file_path = filedialog.askopenfilename(
            filetypes=[("Excel Files", "*.xlsx")],
            title=_('arc_import_title', 'Import Arc Analysis Results'),
            initialdir=init_dir
        )
        if not file_path: return
        
        try:
            import openpyxl
            wb = openpyxl.load_workbook(file_path)
            ws = wb.active
            
            # Patikriname pagrindines antraštes, kad įsitikintume, jog tai teisingas failas (supporting both lt/en)
            expected_headers = [
                ["lanko indeksas", "arc index"], 
                ["temperatūra", "temperature"], 
                ["varža r", "resistance r"], 
                ["cpe rodiklis n", "cpe exponent n"]
            ]
            
            headers = [ws.cell(row=1, column=col_idx).value for col_idx in range(1, len(expected_headers) + 1)]
            headers_clean = [str(h).strip().lower() if h else "" for h in headers]
            
            valid = True
            for exp_list in expected_headers:
                if not any(el in headers_clean for el in exp_list):
                    valid = False
                    break
            
            if not valid:
                messagebox.showerror(_('msg_error', 'Error'), _('arc_import_structure_error', 'Invalid file structure! Could not find required arc columns.'))
                return
            
            # Klausiame ar nori išvalyti ar prirašyti
            if app.locked_arc_data:
                ans = messagebox.askyesnocancel(
                    _('arc_import_dialog_title', 'Import Data'), 
                    _('arc_import_confirm_clear', 
                      "You already have locked arcs. Do you want to clear them before importing?\n"
                      "Selecting 'Yes' will clear existing arcs.\n"
                      "Selecting 'No' will append the imported arcs.\n"
                      "Selecting 'Cancel' will abort the action."),
                    parent=sw
                )
                if ans is None: # Cancel
                    return
                elif ans is True: # Išvalyti
                    for m in app.locked_arc_artists:
                        try: m.remove()
                        except: pass
                    app.locked_arc_artists.clear()
                    app.locked_arc_data.clear()
            
            imported_by_arc = {}
            row_idx = 3
            while True:
                arc_idx_val = ws.cell(row=row_idx, column=1).value
                if arc_idx_val is None:
                    break
                
                temp_val = ws.cell(row=row_idx, column=2).value
                r_fit_val = ws.cell(row=row_idx, column=3).value
                n_cpe_val = ws.cell(row=row_idx, column=4).value
                q_cpe_val = ws.cell(row=row_idx, column=5).value
                c_eq_val = ws.cell(row=row_idx, column=6).value
                x_int1_val = ws.cell(row=row_idx, column=7).value
                x_int2_val = ws.cell(row=row_idx, column=8).value
                chi_square_val = ws.cell(row=row_idx, column=9).value
                
                if chi_square_val is not None:
                    try: chi_square_val = float(chi_square_val)
                    except: chi_square_val = 0.0
                else:
                    chi_square_val = 0.0
                
                if any(v is None for v in [temp_val, r_fit_val, n_cpe_val, q_cpe_val, c_eq_val, x_int1_val, x_int2_val]):
                    row_idx += 1
                    continue
                
                try:
                    temp = float(temp_val)
                    r_fit = float(r_fit_val)
                    n_cpe = float(n_cpe_val)
                    q_cpe = float(q_cpe_val)
                    c_eq = float(c_eq_val)
                    x_int1 = float(x_int1_val)
                    x_int2 = float(x_int2_val)
                except ValueError:
                    row_idx += 1
                    continue
                
                # Rekonstruojame xc, yc, R_circ2 geometriją
                R_fit = r_fit
                phi = (np.pi / 2.0) * (1.0 - n_cpe)
                xc = (x_int1 + x_int2) / 2.0
                if np.cos(phi) != 0:
                    R_circ = R_fit / (2.0 * np.cos(phi))
                    yc = - (R_fit / 2.0) * np.tan(phi)
                else:
                    R_circ = R_fit / 2.0
                    yc = 0.0
                R_circ2 = R_circ**2
                
                x_arc = np.linspace(x_int1, x_int2, 100)
                y_arc = yc + np.sqrt(np.clip(R_circ2 - (x_arc - xc)**2, 0, None))
                
                fit_data = {
                    'temp': temp,
                    'r_fit': r_fit,
                    'n_cpe': n_cpe,
                    'q_cpe': q_cpe,
                    'c_eq': c_eq,
                    'x_int1': x_int1,
                    'x_int2': x_int2,
                    'fit_x': x_arc,
                    'fit_y': y_arc,
                    'chi_square': chi_square_val
                }
                
                arc_key = str(arc_idx_val)
                if arc_key not in imported_by_arc:
                    imported_by_arc[arc_key] = []
                imported_by_arc[arc_key].append(fit_data)
                
                row_idx += 1
            
            for arc_key, temp_fits in imported_by_arc.items():
                app.locked_arc_data.append({
                    'temp_fits': temp_fits,
                    'markers': []
                })
            
            selected_val = temp_var.get()
            if selected_val == _('arc_all_selected', 'All Selected'):
                active_selected = [t for t, v in app.vars.items() if v.get()]
                plot_temps = [str_to_orig[f"{t:g}" if isinstance(t, (int, float)) else str(t)] for t in active_selected]
            else:
                plot_temps = [str_to_orig[selected_val]]
            
            redraw_plot(plot_temps)
            update_sidebar_text()
            
            messagebox.showinfo(_('msg_success', 'Success'), _('arc_import_success_msg', 'Successfully imported {} arcs.').format(len(imported_by_arc)))
        except Exception as e:
            messagebox.showerror(_('msg_error', 'Error'), _('arc_import_error_msg', 'Failed to read Excel file:\n{}').format(e))
 
    # Klavišų paspaudimo įvykiai
    def on_key_press(event):
        if event.key == 's' or event.key == 'S':
            lock_current_arc()
        elif event.key == 'c' or event.key == 'C':
            clear_locked_arcs()
 
    fig.canvas.mpl_connect('key_press_event', on_key_press)
    
    # Priverčiame gauti fokusą klaviatūrai
    canvas_widget = FigureCanvasTkAgg(fig, master=plot_frame)
    canvas_widget.get_tk_widget().bind("<Enter>", lambda e: canvas_widget.get_tk_widget().focus_set())
    
    # Sidebar mygtukai su nuostabiu stiliumi
    btn_lock = tk.Button(btn_frame, text=_('arc_lock_btn', "🔒 Lock Arc (S)"), command=lock_current_arc,
                         bg="#2E7D32", fg="white", font=('Arial', 10, 'bold'), height=2, relief="raised", bd=3)
    btn_lock.pack(fill="x", pady=5)
    
    btn_clear = tk.Button(btn_frame, text=_('arc_clear_btn', "🧹 Clear Locked (C)"), command=clear_locked_arcs,
                          bg="#C62828", fg="white", font=('Arial', 10, 'bold'), height=2, relief="raised", bd=3)
    btn_clear.pack(fill="x", pady=5)
    
    btn_import = tk.Button(sidebar, text=_('arc_import_btn', "📂 Import from Excel"), command=import_locked_arcs_from_excel,
                           bg="#F57C00", fg="white", font=('Arial', 10, 'bold'), height=2, relief="raised", bd=3)
    btn_import.pack(side=tk.BOTTOM, fill="x", padx=15, pady=(5, 15))
    
    btn_excel = tk.Button(sidebar, text=_('arc_export_btn', "📊 Export to Excel"), command=export_locked_arcs_to_excel,
                          bg="#1565C0", fg="white", font=('Arial', 10, 'bold'), height=2, relief="raised", bd=3)
    btn_excel.pack(side=tk.BOTTOM, fill="x", padx=15, pady=5)

    # Toolbaras ir Canvas kairėje
    tb_frame = tk.Frame(plot_frame)
    tb_frame.pack(side=tk.BOTTOM, fill=tk.X)
    
    canvas = FigureCanvasTkAgg(fig, master=plot_frame)
    canvas.get_tk_widget().pack(fill=tk.BOTH, expand=True)
    
    toolbar = NavigationToolbar2Tk(canvas, tb_frame)
    toolbar.update()
    
    fig.subplots_adjust(left=0.07, bottom=0.073, right=0.954, top=0.905, wspace=0.197, hspace=0.183)
    sw.update()
    canvas.draw()
    
    # Vieno pikselio "refresh" triukas
    def force_refresh():
        try:
            geom = sw.geometry()
            parts = geom.split('+')
            wh = parts[0].split('x')
            w_new = int(wh[0]) + 1
            h_new = int(wh[1])
            sw.geometry(f"{w_new}x{h_new}+{parts[1]}+{parts[2]}")
        except: pass
        
    sw.after(200, force_refresh)
